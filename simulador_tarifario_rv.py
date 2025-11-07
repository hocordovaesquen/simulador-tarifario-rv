"""
🚀 SIMULADOR TARIFARIO RV - VERSIÓN PROFESIONAL
================================================
Consultoría en Estructuras Tarifarias para Bolsas de Valores

✅ Lee datos REALES de "A.3 Negociación" (Ingreso Actual vs Propuesta)
✅ Permite editar tramos tarifarios de "1. Parametros"
✅ Simula impacto por broker, producto y país
✅ Análisis "what-if" en tiempo real

Autor: Consultor Especialista en Bolsas de Valores
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from io import BytesIO
import openpyxl

# ==================== CONFIGURACIÓN ====================
st.set_page_config(
    page_title="Simulador Tarifario RV Pro",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==================== ESTILOS ====================
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        background: linear-gradient(120deg, #1e3a8a 0%, #3b82f6 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        padding: 1rem 0;
        margin-bottom: 1rem;
    }
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem;
        border-radius: 10px;
        color: white;
        text-align: center;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .insight-box {
        background-color: #f0f9ff;
        border-left: 4px solid #3b82f6;
        padding: 1rem;
        margin: 1rem 0;
        border-radius: 4px;
    }
</style>
""", unsafe_allow_html=True)

# ==================== FUNCIONES DE CARGA ====================

def limpiar_numero(valor):
    """Limpia y convierte valores a float"""
    if pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    try:
        valor_str = str(valor).replace('$', '').replace(',', '').strip()
        return float(valor_str) if valor_str else 0.0
    except:
        return 0.0

@st.cache_data
def cargar_datos_negociacion(archivo):
    """
    Carga datos desde A.3 Negociación.
    La hoja tiene 3 secciones horizontales por país:
    - CHILE: Cols C-T (3-20)
    - COLOMBIA: Cols AH+ (34+) 
    - PERÚ: Cols AP+ (42+)
    """
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb['A.3 Negociación']
        
        datos = []
        
        # SECCIÓN 1: CHILE (Columnas 3-20)
        for row_idx in range(10, 100):
            broker = ws.cell(row=row_idx, column=3).value
            if not broker or str(broker).strip() == "":
                break
            
            monto = limpiar_numero(ws.cell(row=row_idx, column=5).value)
            total_real = limpiar_numero(ws.cell(row=row_idx, column=6).value)
            total_prop = limpiar_numero(ws.cell(row=row_idx, column=7).value)
            acceso_real = limpiar_numero(ws.cell(row=row_idx, column=9).value)
            acceso_prop = limpiar_numero(ws.cell(row=row_idx, column=10).value)
            trans_real = limpiar_numero(ws.cell(row=row_idx, column=12).value)
            trans_prop = limpiar_numero(ws.cell(row=row_idx, column=13).value)
            codigos_real = limpiar_numero(ws.cell(row=row_idx, column=15).value)
            codigos_prop = limpiar_numero(ws.cell(row=row_idx, column=16).value)
            
            datos.append({
                'Broker': str(broker).strip(),
                'Pais': 'Chile',
                'Monto_USD': monto,
                'Acceso_Real': acceso_real,
                'Acceso_Propuesta': acceso_prop,
                'Trans_Real': trans_real,
                'Trans_Propuesta': trans_prop,
                'Codigos_Real': codigos_real,
                'Codigos_Propuesta': codigos_prop,
                'Total_Real': total_real,
                'Total_Propuesta': total_prop
            })
        
        # SECCIÓN 2: COLOMBIA (Columnas 34+, estructura similar)
        for row_idx in range(10, 100):
            broker = ws.cell(row=row_idx, column=34).value
            if not broker or str(broker).strip() == "":
                break
            
            # Ajustar offset de columnas (34 es el inicio, restar 31 para obtener offset relativo)
            offset = 31
            monto = limpiar_numero(ws.cell(row=row_idx, column=36).value)
            total_real = limpiar_numero(ws.cell(row=row_idx, column=37).value)
            total_prop = limpiar_numero(ws.cell(row=row_idx, column=38).value)
            acceso_real = limpiar_numero(ws.cell(row=row_idx, column=40).value)
            acceso_prop = limpiar_numero(ws.cell(row=row_idx, column=41).value)
            trans_real = limpiar_numero(ws.cell(row=row_idx, column=43).value)
            trans_prop = limpiar_numero(ws.cell(row=row_idx, column=44).value)
            codigos_real = limpiar_numero(ws.cell(row=row_idx, column=46).value)
            codigos_prop = limpiar_numero(ws.cell(row=row_idx, column=47).value)
            
            datos.append({
                'Broker': str(broker).strip(),
                'Pais': 'Colombia',
                'Monto_USD': monto,
                'Acceso_Real': acceso_real,
                'Acceso_Propuesta': acceso_prop,
                'Trans_Real': trans_real,
                'Trans_Propuesta': trans_prop,
                'Codigos_Real': codigos_real,
                'Codigos_Propuesta': codigos_prop,
                'Total_Real': total_real,
                'Total_Propuesta': total_prop
            })
        
        # SECCIÓN 3: PERÚ (Columnas 42+)
        for row_idx in range(10, 100):
            broker = ws.cell(row=row_idx, column=42).value
            if not broker or str(broker).strip() == "":
                break
            
            # Offset para Perú
            offset_peru = 39
            monto = limpiar_numero(ws.cell(row=row_idx, column=44).value)
            total_real = limpiar_numero(ws.cell(row=row_idx, column=45).value)
            total_prop = limpiar_numero(ws.cell(row=row_idx, column=46).value)
            acceso_real = limpiar_numero(ws.cell(row=row_idx, column=48).value)
            acceso_prop = limpiar_numero(ws.cell(row=row_idx, column=49).value)
            trans_real = limpiar_numero(ws.cell(row=row_idx, column=51).value)
            trans_prop = limpiar_numero(ws.cell(row=row_idx, column=52).value)
            codigos_real = limpiar_numero(ws.cell(row=row_idx, column=54).value)
            codigos_prop = limpiar_numero(ws.cell(row=row_idx, column=55).value)
            
            datos.append({
                'Broker': str(broker).strip(),
                'Pais': 'Perú',
                'Monto_USD': monto,
                'Acceso_Real': acceso_real,
                'Acceso_Propuesta': acceso_prop,
                'Trans_Real': trans_real,
                'Trans_Propuesta': trans_prop,
                'Codigos_Real': codigos_real,
                'Codigos_Propuesta': codigos_prop,
                'Total_Real': total_real,
                'Total_Propuesta': total_prop
            })
        
        df = pd.DataFrame(datos)
        
        # Calcular diferencias y variaciones
        df['Dif_Acceso'] = df['Acceso_Propuesta'] - df['Acceso_Real']
        df['Dif_Trans'] = df['Trans_Propuesta'] - df['Trans_Real']
        df['Dif_Codigos'] = df['Codigos_Propuesta'] - df['Codigos_Real']
        df['Dif_Total'] = df['Total_Propuesta'] - df['Total_Real']
        
        df['Var_Total_%'] = df.apply(
            lambda x: ((x['Dif_Total'] / x['Total_Real']) * 100) if x['Total_Real'] > 0 else 0, 
            axis=1
        )
        
        # BPS
        df['BPS_Real'] = df.apply(
            lambda x: (x['Total_Real'] / x['Monto_USD'] * 10000) if x['Monto_USD'] > 0 else 0,
            axis=1
        )
        df['BPS_Propuesta'] = df.apply(
            lambda x: (x['Total_Propuesta'] / x['Monto_USD'] * 10000) if x['Monto_USD'] > 0 else 0,
            axis=1
        )
        
        st.success(f"✅ {len(df)} brokers cargados desde A.3 Negociación")
        return df
        
    except Exception as e:
        st.error(f"❌ Error al cargar datos: {str(e)}")
        return None

@st.cache_data
def cargar_parametros(archivo):
    """
    Carga parámetros tarifarios desde "1. Parametros"
    Estructura:
    - Filas 100-103: Tramos de ACCESO
    - Filas 139-142: Tramos de TRANSACCIÓN
    
    Columnas por país:
    - Colombia: T-W (20-23): Min, Max, Variable%, Fija$
    - Perú: X-AA (24-27): Min, Max, Variable%, Fija$
    - Chile: AB-AE (28-31): Min, Max, Variable%, Fija$
    """
    parametros = {
        'Acceso': {'Colombia': [], 'Peru': [], 'Perú': [], 'Chile': []},
        'Transaccion': {'Colombia': [], 'Peru': [], 'Perú': [], 'Chile': []}
    }
    
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb['1. Parametros']
        
        # Mapeo de filas a productos
        rangos = {
            'Acceso': (100, 103),      # Filas 100-103
            'Transaccion': (139, 142)   # Filas 139-142
        }
        
        # Mapeo de columnas por país (basado en índice 1-based)
        # T=20, U=21, V=22, W=23
        # X=24, Y=25, Z=26, AA=27
        # AB=28, AC=29, AD=30, AE=31
        columnas_pais = {
            'Colombia': (20, 21, 22, 23),  # T, U, V, W
            'Peru': (24, 25, 26, 27),       # X, Y, Z, AA
            'Chile': (28, 29, 30, 31)       # AB, AC, AD, AE
        }
        
        for producto, (fila_inicio, fila_fin) in rangos.items():
            for fila in range(fila_inicio, fila_fin + 1):
                for pais, (col_min, col_max, col_var, col_fija) in columnas_pais.items():
                    try:
                        min_val = limpiar_numero(ws.cell(row=fila, column=col_min).value)
                        max_val = limpiar_numero(ws.cell(row=fila, column=col_max).value)
                        var_val = limpiar_numero(ws.cell(row=fila, column=col_var).value)
                        fija_val = limpiar_numero(ws.cell(row=fila, column=col_fija).value)
                        
                        # Manejar valores infinitos
                        if max_val > 1e15:
                            max_val = float('inf')
                        
                        # Solo agregar si tiene valores válidos
                        if min_val >= 0 and (max_val > 0 or var_val > 0 or fija_val > 0):
                            tramo = {
                                'min': min_val,
                                'max': max_val,
                                'var': var_val,
                                'fija': fija_val
                            }
                            parametros[producto][pais].append(tramo)
                            # También agregar con acento para compatibilidad
                            if pais == 'Peru':
                                parametros[producto]['Perú'].append(tramo)
                    except:
                        continue
        
        # Valores por defecto si no hay datos
        for producto in ['Acceso', 'Transaccion']:
            for pais in ['Colombia', 'Peru', 'Perú', 'Chile']:
                if not parametros[producto][pais]:
                    parametros[producto][pais] = [
                        {'min': 0, 'max': 5_000_000, 'var': 0, 'fija': 500},
                        {'min': 5_000_001, 'max': 15_000_000, 'var': 0, 'fija': 1500},
                        {'min': 15_000_001, 'max': float('inf'), 'var': 0, 'fija': 3000}
                    ]
        
        st.success(f"✅ Parámetros tarifarios cargados")
        return parametros
        
    except Exception as e:
        st.warning(f"⚠️ Usando parámetros por defecto")
        # Valores por defecto
        for producto in ['Acceso', 'Transaccion']:
            for pais in ['Colombia', 'Peru', 'Perú', 'Chile']:
                parametros[producto][pais] = [
                    {'min': 0, 'max': 5_000_000, 'var': 0, 'fija': 500},
                    {'min': 5_000_001, 'max': 15_000_000, 'var': 0, 'fija': 1500},
                    {'min': 15_000_001, 'max': float('inf'), 'var': 0, 'fija': 3000}
                ]
        return parametros

# ==================== FUNCIONES DE CÁLCULO ====================

def calcular_ingreso_tramo(monto, tramos):
    """Calcula ingreso según tramos tarifarios"""
    if not tramos or monto == 0:
        return 0.0
    
    for tramo in tramos:
        if tramo['min'] <= monto < tramo['max'] or \
           (tramo['max'] == float('inf') and monto >= tramo['min']):
            return (monto * tramo['var'] / 100) + tramo['fija']
    
    # Si no encaja en ningún tramo, usar el último
    ultimo = tramos[-1]
    return (monto * ultimo['var'] / 100) + ultimo['fija']

def simular_con_nuevos_parametros(df_datos, parametros):
    """
    Simula nuevos ingresos basados en parámetros editados
    """
    resultados = df_datos.copy()
    
    # Simular nuevos valores
    for idx, row in resultados.iterrows():
        pais = row['Pais']
        monto = row['Monto_USD']
        
        # Normalizar nombre del país
        pais_key = pais
        if pais not in ['Colombia', 'Peru', 'Perú', 'Chile']:
            pais_key = 'Chile'  # Por defecto
        
        # Calcular nuevos ingresos
        tramos_acceso = parametros['Acceso'].get(pais_key, [])
        tramos_trans = parametros['Transaccion'].get(pais_key, [])
        
        nuevo_acceso = calcular_ingreso_tramo(monto, tramos_acceso)
        nuevo_trans = calcular_ingreso_tramo(monto, tramos_trans)
        
        resultados.at[idx, 'Acceso_Simulado'] = nuevo_acceso
        resultados.at[idx, 'Trans_Simulado'] = nuevo_trans
        resultados.at[idx, 'Total_Simulado'] = nuevo_acceso + nuevo_trans + row['Codigos_Propuesta']
    
    # Recalcular diferencias
    resultados['Dif_Simulado'] = resultados['Total_Simulado'] - resultados['Total_Real']
    resultados['Var_Simulado_%'] = resultados.apply(
        lambda x: ((x['Dif_Simulado'] / x['Total_Real']) * 100) if x['Total_Real'] > 0 else 0,
        axis=1
    )
    
    return resultados

# ==================== VISUALIZACIONES ====================

def crear_grafico_comparativo(df):
    """Gráfico comparativo Real vs Propuesta vs Simulado"""
    total_real = df['Total_Real'].sum()
    total_prop = df['Total_Propuesta'].sum()
    total_sim = df['Total_Simulado'].sum() if 'Total_Simulado' in df.columns else total_prop
    
    fig = go.Figure(data=[
        go.Bar(name='Real Actual', x=['Total'], y=[total_real], 
               marker_color='#ef4444',
               text=[f'${total_real/1e6:.2f}M'], textposition='outside'),
        go.Bar(name='Propuesta Excel', x=['Total'], y=[total_prop],
               marker_color='#3b82f6',
               text=[f'${total_prop/1e6:.2f}M'], textposition='outside'),
        go.Bar(name='Simulado', x=['Total'], y=[total_sim],
               marker_color='#22c55e',
               text=[f'${total_sim/1e6:.2f}M'], textposition='outside')
    ])
    
    fig.update_layout(
        title='<b>Comparación de Ingresos Totales</b>',
        barmode='group',
        height=400,
        template='plotly_white',
        showlegend=True
    )
    return fig

def crear_grafico_por_pais(df):
    """Gráfico de ingresos por país"""
    df_pais = df.groupby('Pais').agg({
        'Total_Real': 'sum',
        'Total_Propuesta': 'sum'
    }).reset_index()
    
    fig = go.Figure(data=[
        go.Bar(name='Real', x=df_pais['Pais'], y=df_pais['Total_Real'],
               marker_color='#ef4444'),
        go.Bar(name='Propuesta', x=df_pais['Pais'], y=df_pais['Total_Propuesta'],
               marker_color='#22c55e')
    ])
    
    fig.update_layout(
        title='<b>Ingresos por País</b>',
        barmode='group',
        height=400,
        template='plotly_white'
    )
    return fig

def crear_grafico_por_producto(df):
    """Gráfico de ingresos por producto"""
    productos = {
        'Acceso': [df['Acceso_Real'].sum(), df['Acceso_Propuesta'].sum()],
        'Transacción': [df['Trans_Real'].sum(), df['Trans_Propuesta'].sum()],
        'Códigos': [df['Codigos_Real'].sum(), df['Codigos_Propuesta'].sum()]
    }
    
    fig = go.Figure(data=[
        go.Bar(name='Real', x=list(productos.keys()), 
               y=[v[0] for v in productos.values()],
               marker_color='#ef4444'),
        go.Bar(name='Propuesta', x=list(productos.keys()), 
               y=[v[1] for v in productos.values()],
               marker_color='#22c55e')
    ])
    
    fig.update_layout(
        title='<b>Ingresos por Producto</b>',
        barmode='group',
        height=400,
        template='plotly_white'
    )
    return fig

def crear_grafico_top_brokers(df, n=10):
    """Top N brokers por diferencia en ingresos"""
    df_sorted = df.nlargest(n, 'Dif_Total')
    
    fig = go.Figure(data=[
        go.Bar(
            x=df_sorted['Dif_Total'],
            y=df_sorted['Broker'],
            orientation='h',
            marker_color=df_sorted['Dif_Total'].apply(
                lambda x: '#22c55e' if x > 0 else '#ef4444'
            ),
            text=df_sorted['Dif_Total'].apply(lambda x: f'${x/1e3:.1f}K'),
            textposition='outside'
        )
    ])
    
    fig.update_layout(
        title=f'<b>Top {n} Brokers por Diferencia (Propuesta - Real)</b>',
        height=400,
        template='plotly_white',
        xaxis_title='Diferencia ($)',
        yaxis_title=''
    )
    return fig

# ==================== MAIN ====================

def main():
    st.markdown('<h1 class="main-header">📊 SIMULADOR TARIFARIO RV - VERSIÓN PROFESIONAL</h1>', 
                unsafe_allow_html=True)
    
    st.markdown("""
    <div class="insight-box">
    <b>🎯 Funcionalidad:</b> Este simulador te permite comparar los <b>Ingresos Actuales</b> 
    contra la <b>Propuesta Tarifaria</b> del Excel, y además <b>editar los tramos</b> en tiempo real 
    para ver el impacto inmediato en tus ingresos por broker, producto y país.
    </div>
    """, unsafe_allow_html=True)
    
    # ==================== SIDEBAR ====================
    with st.sidebar:
        st.markdown("## ⚙️ Configuración")
        
        archivo = st.file_uploader("📁 Cargar Excel de Tarifas", type=['xlsx'])
        
        if not archivo:
            st.info("👆 Sube tu archivo Excel")
            st.markdown("""
            **Estructura requerida:**
            - Hoja: `A.3 Negociación` (datos por broker)
            - Hoja: `1. Parametros` (tramos tarifarios)
            """)
            st.stop()
        
        # Cargar datos
        with st.spinner("📊 Cargando datos..."):
            df_datos = cargar_datos_negociacion(archivo)
            if df_datos is None:
                st.stop()
            
            # Cargar parámetros en session_state
            if 'parametros' not in st.session_state:
                st.session_state.parametros = cargar_parametros(archivo)
        
        st.markdown("---")
        
        # Filtros
        st.markdown("### 🔍 Filtros")
        paises = ['Todos'] + sorted(df_datos['Pais'].unique().tolist())
        pais_filtro = st.selectbox("🌎 País", paises, key='filtro_pais')
        
        st.markdown("---")
        
        # Editor de parámetros
        st.markdown("### ✏️ Editor de Tramos")
        
        editar = st.checkbox("🔓 Habilitar Edición", value=False)
        
        if editar:
            st.markdown("#### Configurar Tramos:")
            
            pais_edit = st.selectbox(
                "País",
                ['Colombia', 'Peru', 'Chile'],
                key='pais_edit'
            )
            
            producto_edit = st.selectbox(
                "Producto",
                ['Acceso', 'Transaccion'],
                key='producto_edit'
            )
            
            st.markdown(f"**Tramos para {pais_edit} - {producto_edit}:**")
            
            # Obtener tramos actuales
            tramos_actuales = st.session_state.parametros[producto_edit][pais_edit]
            
            # Editar cada tramo
            tramos_nuevos = []
            for i, tramo in enumerate(tramos_actuales):
                with st.expander(f"📊 Tramo {i+1}", expanded=(i<2)):
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        min_val = st.number_input(
                            "Mínimo (USD)",
                            value=float(tramo['min']),
                            step=100000.0,
                            key=f"min_{pais_edit}_{producto_edit}_{i}",
                            format="%.0f"
                        )
                        
                        var_val = st.number_input(
                            "Variable %",
                            value=float(tramo['var']),
                            step=0.01,
                            key=f"var_{pais_edit}_{producto_edit}_{i}",
                            format="%.4f"
                        )
                    
                    with col2:
                        max_display = tramo['max'] if tramo['max'] != float('inf') else 999999999999
                        max_val = st.number_input(
                            "Máximo (USD)",
                            value=float(max_display),
                            step=100000.0,
                            key=f"max_{pais_edit}_{producto_edit}_{i}",
                            format="%.0f"
                        )
                        if max_val > 1e12:
                            max_val = float('inf')
                        
                        fija_val = st.number_input(
                            "Fija (USD)",
                            value=float(tramo['fija']),
                            step=100.0,
                            key=f"fija_{pais_edit}_{producto_edit}_{i}",
                            format="%.2f"
                        )
                    
                    tramos_nuevos.append({
                        'min': min_val,
                        'max': max_val,
                        'var': var_val,
                        'fija': fija_val
                    })
            
            # Actualizar parámetros
            st.session_state.parametros[producto_edit][pais_edit] = tramos_nuevos
            # También actualizar con acento
            if pais_edit == 'Peru':
                st.session_state.parametros[producto_edit]['Perú'] = tramos_nuevos
            
            if st.button("🔄 Aplicar y Recalcular", use_container_width=True):
                st.success("✅ Tramos actualizados!")
                st.rerun()
    
    # ==================== CONTENIDO PRINCIPAL ====================
    
    # Filtrar datos
    if pais_filtro == 'Todos':
        df_filtrado = df_datos
    else:
        df_filtrado = df_datos[df_datos['Pais'] == pais_filtro]
    
    # Simular con parámetros actuales
    with st.spinner("🎯 Simulando..."):
        df_resultados = simular_con_nuevos_parametros(df_filtrado, st.session_state.parametros)
    
    # ==================== KPIs ====================
    st.markdown("### 💰 Métricas Principales")
    
    col1, col2, col3, col4 = st.columns(4)
    
    total_monto = df_resultados['Monto_USD'].sum()
    total_real = df_resultados['Total_Real'].sum()
    total_prop = df_resultados['Total_Propuesta'].sum()
    total_sim = df_resultados['Total_Simulado'].sum()
    
    dif_prop = total_prop - total_real
    var_prop = (dif_prop / total_real * 100) if total_real > 0 else 0
    
    dif_sim = total_sim - total_real
    var_sim = (dif_sim / total_real * 100) if total_real > 0 else 0
    
    with col1:
        st.metric(
            "💵 Monto Negociado",
            f"${total_monto/1e6:.1f}M",
            f"{len(df_resultados)} brokers"
        )
    
    with col2:
        st.metric(
            "📊 Ingreso Real Actual",
            f"${total_real/1e6:.2f}M",
            f"{(total_real/total_monto)*10000:.1f} bps"
        )
    
    with col3:
        st.metric(
            "🎯 Propuesta Excel",
            f"${total_prop/1e6:.2f}M",
            f"{var_prop:+.1f}%",
            delta_color="normal" if var_prop >= 0 else "inverse"
        )
    
    with col4:
        st.metric(
            "⚡ Simulado (Editado)",
            f"${total_sim/1e6:.2f}M",
            f"{var_sim:+.1f}%",
            delta_color="normal" if var_sim >= 0 else "inverse"
        )
    
    st.markdown("---")
    
    # ==================== INSIGHTS ====================
    st.markdown("### 💡 Insights Clave")
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Brokers con mayor ganancia
        top_ganadores = df_resultados.nlargest(3, 'Dif_Total')
        st.markdown("**🟢 Top 3 Mayor Ganancia (Propuesta vs Real):**")
        for idx, row in top_ganadores.iterrows():
            st.markdown(f"- **{row['Broker'][:30]}**: +${row['Dif_Total']/1e3:.1f}K ({row['Var_Total_%']:+.1f}%)")
    
    with col2:
        # Brokers con mayor pérdida
        top_perdedores = df_resultados.nsmallest(3, 'Dif_Total')
        st.markdown("**🔴 Top 3 Mayor Pérdida (Propuesta vs Real):**")
        for idx, row in top_perdedores.iterrows():
            st.markdown(f"- **{row['Broker'][:30]}**: ${row['Dif_Total']/1e3:.1f}K ({row['Var_Total_%']:+.1f}%)")
    
    st.markdown("---")
    
    # ==================== GRÁFICOS ====================
    st.markdown("### 📈 Análisis Visual")
    
    tab1, tab2, tab3, tab4 = st.tabs([
        "📊 Comparativo Total",
        "🌎 Por País",
        "📦 Por Producto",
        "🏆 Top Brokers"
    ])
    
    with tab1:
        st.plotly_chart(crear_grafico_comparativo(df_resultados), use_container_width=True)
    
    with tab2:
        st.plotly_chart(crear_grafico_por_pais(df_resultados), use_container_width=True)
    
    with tab3:
        st.plotly_chart(crear_grafico_por_producto(df_resultados), use_container_width=True)
    
    with tab4:
        n_top = st.slider("Número de brokers a mostrar", 5, 20, 10)
        st.plotly_chart(crear_grafico_top_brokers(df_resultados, n_top), use_container_width=True)
    
    st.markdown("---")
    
    # ==================== TABLA DETALLADA ====================
    st.markdown("### 📋 Detalle por Broker")
    
    # Preparar datos para display
    df_display = df_resultados.copy()
    
    # Formatear columnas
    df_display['Monto_USD'] = df_display['Monto_USD'].apply(lambda x: f"${x:,.0f}")
    df_display['Total_Real'] = df_display['Total_Real'].apply(lambda x: f"${x:,.2f}")
    df_display['Total_Propuesta'] = df_display['Total_Propuesta'].apply(lambda x: f"${x:,.2f}")
    df_display['Dif_Total'] = df_display['Dif_Total'].apply(lambda x: f"${x:,.2f}")
    df_display['Var_Total_%'] = df_display['Var_Total_%'].apply(lambda x: f"{x:+.1f}%")
    df_display['BPS_Real'] = df_display['BPS_Real'].apply(lambda x: f"{x:.1f}")
    df_display['BPS_Propuesta'] = df_display['BPS_Propuesta'].apply(lambda x: f"{x:.1f}")
    
    # Seleccionar columnas a mostrar
    columnas_mostrar = [
        'Broker', 'Pais', 'Monto_USD', 
        'Total_Real', 'BPS_Real',
        'Total_Propuesta', 'BPS_Propuesta',
        'Dif_Total', 'Var_Total_%'
    ]
    
    st.dataframe(
        df_display[columnas_mostrar].sort_values('Var_Total_%', ascending=False),
        use_container_width=True,
        height=500
    )
    
    # ==================== EXPORTAR ====================
    st.markdown("---")
    st.markdown("### 💾 Exportar Resultados")
    
    col1, col2 = st.columns(2)
    
    with col1:
        csv = df_resultados.to_csv(index=False).encode('utf-8')
        st.download_button(
            "📥 Descargar CSV Completo",
            csv,
            "simulacion_tarifaria_rv.csv",
            "text/csv",
            use_container_width=True
        )
    
    with col2:
        # Resumen ejecutivo
        resumen = pd.DataFrame({
            'Métrica': ['Monto Total', 'Ingreso Real', 'Ingreso Propuesta', 'Diferencia', 'Variación %'],
            'Valor': [
                f"${total_monto:,.0f}",
                f"${total_real:,.2f}",
                f"${total_prop:,.2f}",
                f"${dif_prop:,.2f}",
                f"{var_prop:.2f}%"
            ]
        })
        csv_resumen = resumen.to_csv(index=False).encode('utf-8')
        st.download_button(
            "📊 Descargar Resumen Ejecutivo",
            csv_resumen,
            "resumen_ejecutivo.csv",
            "text/csv",
            use_container_width=True
        )

# ==================== EJECUTAR ====================
if __name__ == "__main__":
    main()
