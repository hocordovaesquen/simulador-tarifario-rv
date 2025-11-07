# app.py
# ============================================================================
# Simulador Tarifario RV — v8 (Excel-driven, robust + safe save)
# ----------------------------------------------------------------------------
# - Edita tramos en "1. Parametros" (columna R+: T:AE, filas 117–160) y escribe
#   directamente en el .xlsx (manejo seguro de celdas combinadas).
# - Recalcula con xlcalculator si está disponible; si no, usa valores cached
#   (data_only=True); último recurso: fallback sin romper.
# - Lee "A.3 Negociación" fila por fila, detectando columnas por encabezados
#   y, si no hay "Ingreso Total", suma Acceso + Transacción + Perfiles.
# - Asigna País por bloques BCS/BVL/BVC.
# - Al descargar el Excel, elimina imágenes/charts para evitar errores de PIL/EMF.
#
# Requisitos:
#   pip install -r requirements.txt
#   streamlit run app.py
# ============================================================================

import io
import math
import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go
from io import BytesIO
from typing import Dict, List, Tuple
import openpyxl
from openpyxl.cell.cell import MergedCell

st.set_page_config(page_title="Simulador Tarifario RV – Excel-driven (v8)", page_icon="📊", layout="wide")

# ---- Intento de importar xlcalculator ----
try:
    from xlcalculator import ModelCompiler, Model
    XL_OK = True
except Exception:
    XL_OK = False

# -----------------------------
# Helpers
# -----------------------------
def is_num(x):
    return isinstance(x, (int, float)) and not (isinstance(x, float) and math.isnan(x))

def safe_float(x, default=0.0):
    try:
        if x is None or (isinstance(x, float) and math.isnan(x)):
            return default
        if isinstance(x, str):
            x = x.replace("$","").replace(",","").replace("\\xa0","").strip()
        return float(x)
    except:
        return default

def calc_bps(ingreso, monto):
    if monto <= 0:
        return 0.0
    return (ingreso / monto) * 10000.0

def norm(s: str) -> str:
    if s is None: return ""
    s = str(s).strip().lower()
    trans = str.maketrans("áéíóúüñ", "aeiouun")
    return s.translate(trans)

# -----------------------------
# Workbook I/O and safe save
# -----------------------------
PARAM_SHEET = "1. Parametros"
NEG_SHEET   = "A.3 Negociación"
MAP_RIGHT = {"Colombia": ("T","U","V","W"), "Perú": ("X","Y","Z","AA"), "Chile": ("AB","AC","AD","AE")}
ROWS_PANT   = list(range(117, 135))
ROWS_TRX    = list(range(134, 145))
ROWS_DMA    = list(range(146, 160))
DESC_BCS_POS= ("C", 114)

def load_workbook(bytes_data: bytes):
    return openpyxl.load_workbook(io.BytesIO(bytes_data), data_only=False)

def safe_set(ws, coord: str, value):
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        # write to the anchor (top-left) of the merged range
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                ws.cell(row=rng.min_row, column=rng.min_col).value = value
                return
        return
    cell.value = value

def remove_drawings_images(wb):
    """Remove images/charts so openpyxl can save even if the file had EMF/WMF."""
    for ws in wb.worksheets:
        # images
        if hasattr(ws, "_images"):
            ws._images = []
        # charts
        if hasattr(ws, "_charts"):
            ws._charts = []
        # drawings anchors
        if hasattr(ws, "_drawing"):
            ws._drawing = None
    return wb

def safe_save_wb(wb) -> bytes:
    """Try saving workbook. If it fails due to images, strip drawings and retry."""
    out = BytesIO()
    try:
        wb.save(out)
        return out.getvalue()
    except Exception:
        # try removing drawings / images
        try:
            remove_drawings_images(wb)
            out2 = BytesIO()
            wb.save(out2)
            return out2.getvalue()
        except Exception:
            # final fallback: return empty bytes
            return b""

# -----------------------------
# Read / Write parameters
# -----------------------------
def read_block_from_wb(wb, rows, fields=("min","max","bps","fijo")):
    ws = wb[PARAM_SHEET]
    out = {k: [] for k in MAP_RIGHT}
    for pais, (cmin, cmax, cvar, cfijo) in MAP_RIGHT.items():
        for rr in rows:
            mn = ws[f"{cmin}{rr}"].value
            mx = ws[f"{cmax}{rr}"].value
            var = ws[f"{cvar}{rr}"].value
            fijo= ws[f"{cfijo}{rr}"].value
            if any(is_num(v) for v in [mn, mx, var, fijo]):
                out[pais].append({fields[0]: safe_float(mn, 0.0), fields[1]: safe_float(mx, float("inf")),
                                  fields[2]: safe_float(var, 0.0), fields[3]: safe_float(fijo, 0.0)})
    return out

def read_params_from_wb(wb):
    ws = wb[PARAM_SHEET]
    desc = safe_float(ws[f"{DESC_BCS_POS[0]}{DESC_BCS_POS[1]}"].value, 0.15)
    trans = read_block_from_wb(wb, ROWS_TRX, ("min","max","bps","fijo"))
    dma   = read_block_from_wb(wb, ROWS_DMA, ("min","max","bps","fijo"))
    pant  = read_block_from_wb(wb, ROWS_PANT, ("min","max","var","fija"))
    return {"desc_bcs":desc, "transaccion":trans, "dma":dma, "pantallas":pant}

def write_params_to_wb(wb, params: Dict):
    ws = wb[PARAM_SHEET]
    desc = params.get("desc_bcs", None)
    if desc is not None:
        c, r = DESC_BCS_POS
        safe_set(ws, f"{c}{r}", float(desc))
    def write_block(rows: List[int], data_key: str, field_names: Tuple[str,str,str,str]):
        block = params.get(data_key, {})
        for pais, cols in MAP_RIGHT.items():
            cmin, cmax, cvar, cfijo = cols
            lst = block.get(pais, [])
            for i, rr in enumerate(rows):
                payload = lst[i] if i < len(lst) else None
                vmin  = safe_float(payload.get(field_names[0], payload.get("min", 0))) if payload else 0.0
                vmax  = safe_float(payload.get(field_names[1], payload.get("max", 0))) if payload else 0.0
                vvar  = safe_float(payload.get(field_names[2], payload.get("bps", payload.get("var", 0)))) if payload else 0.0
                vfijo = safe_float(payload.get(field_names[3], payload.get("fijo", payload.get("fija", 0)))) if payload else 0.0
                safe_set(ws, f"{cmin}{rr}", vmin)
                safe_set(ws, f"{cmax}{rr}", vmax)
                safe_set(ws, f"{cvar}{rr}", vvar)
                safe_set(ws, f"{cfijo}{rr}", vfijo)
    write_block(ROWS_TRX, "transaccion", ("min","max","bps","fijo"))
    write_block(ROWS_DMA, "dma", ("min","max","bps","fijo"))
    write_block(ROWS_PANT, "pantallas", ("min","max","var","fija"))
    return wb

# -----------------------------
# A.3 Negociación – column detection
# -----------------------------
def find_header_row(ws):
    max_row = min(150, ws.max_row)
    for r in range(1, max_row+1):
        row_norm = [norm(ws.cell(r,c).value) for c in range(1, ws.max_column+1)]
        if any(v.startswith("corredor") or v.startswith("cliente") for v in row_norm):
            return r
    return 9

def build_super_sub(ws, h):
    sup = {c: ws.cell(h, c).value for c in range(1, ws.max_column+1)}
    sub = {c: ws.cell(h+1, c).value for c in range(1, ws.max_column+1)}
    return sup, sub

def locate_columns(ws):
    h = find_header_row(ws)
    sup, sub = build_super_sub(ws, h)
    def has_tokens(text, tokens):
        t = norm(text); return all(tok in t for tok in tokens)
    def find_col(super_tokens: List[str], sub_tokens_any: List[List[str]]):
        # exact match in two-level header
        for subs in sub_tokens_any:
            for c in range(1, ws.max_column+1):
                if has_tokens(sup.get(c,""), super_tokens) and has_tokens(sub.get(c,""), subs):
                    return c
        # combined fallback
        for subs in sub_tokens_any:
            for c in range(1, ws.max_column+1):
                if has_tokens(str(sup.get(c,"")) + " " + str(sub.get(c,"")), super_tokens + subs):
                    return c
        return None
    # basic cols
    cliente = None; monto = None
    for c in range(1, ws.max_column+1):
        s = norm(sup.get(c,"")); t = norm(sub.get(c,""))
        if cliente is None and (("corredor" in s) or ("cliente" in s) or ("corredor" in t) or ("cliente" in t)):
            cliente = c
        if monto is None:
            if (("monto" in s and ("neg" in s or "usd" in s)) or ("monto" in t and ("neg" in t or "usd" in t))):
                monto = c
    # totals
    real_tot = find_col(["ingreso","total"], [["real"],["real "],["real  "]])
    proy_tot = find_col(["ingreso","total"], [["proy"],["proyectado"],["propuesta"],["work"]])
    # components
    acc_r = find_col(["ingreso","acceso"], [["real"]])
    trx_r = find_col(["ingreso","trans"],  [["real"]])
    per_r = find_col(["ingreso","perfil"], [["real"]])
    acc_p = find_col(["ingreso","acceso"], [["proy"],["proyectado"],["propuesta"],["work"]])
    trx_p = find_col(["ingreso","trans"],  [["proy"],["proyectado"],["propuesta"],["work"]])
    per_p = find_col(["ingreso","perfil"], [["proy"],["proyectado"],["propuesta"],["work"]])
    return {"header": h, "cliente": cliente, "monto": monto,
            "real_total": real_tot, "proy_total": proy_tot,
            "acc_real": acc_r, "trx_real": trx_r, "per_real": per_r,
            "acc_proy": acc_p, "trx_proy": trx_p, "per_proy": per_p}

# -----------------------------
# Engines to extract values from A.3
# -----------------------------
def eval_with_model(model, sheet_name, addr):
    if not addr: return 0.0
    try:
        return safe_float(model.evaluate(f"'{sheet_name}'!{addr}"), 0.0)
    except Exception:
        return 0.0

def read_val(ws, addr):
    if not addr: return 0.0
    return safe_float(ws[addr].value, 0.0)

def extract_with_model(wb):
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    mc = ModelCompiler(); model = Model(mc.read_and_parse_archive(bio))
    ws = wb[NEG_SHEET]; cols = locate_columns(ws); h = cols["header"]; ccli = cols["cliente"]
    bolsas_map = {"BCS":"Chile","BVL":"Perú","BVC":"Colombia"}
    rows_cli = []; rows_bol = []; current_pais = None
    for r in range(h+2, ws.max_row+1):
        name = ws.cell(r, ccli).value
        if name is None or str(name).strip()=="": continue
        name_s = str(name).strip()
        def coord(col): return ws.cell(r, col).coordinate if col else None
        if name_s in bolsas_map:
            current_pais = bolsas_map[name_s]
            rt = eval_with_model(model, NEG_SHEET, coord(cols["real_total"] or cols["acc_real"] or cols["trx_real"] or cols["per_real"]))
            pt = eval_with_model(model, NEG_SHEET, coord(cols["proy_total"] or cols["acc_proy"] or cols["trx_proy"] or cols["per_proy"]))
            rows_bol.append({"Pais": current_pais, "Bolsa": name_s, "Real Excel": rt, "Proyectado Excel": pt})
        else:
            monto = eval_with_model(model, NEG_SHEET, coord(cols["monto"]))
            rt = eval_with_model(model, NEG_SHEET, coord(cols["real_total"])) if cols["real_total"] else 0.0
            pt = eval_with_model(model, NEG_SHEET, coord(cols["proy_total"])) if cols["proy_total"] else 0.0
            if rt == 0.0:
                rt = sum(eval_with_model(model, NEG_SHEET, coord(c)) for c in [cols["acc_real"], cols["trx_real"], cols["per_real"]] if c)
            if pt == 0.0:
                pt = sum(eval_with_model(model, NEG_SHEET, coord(c)) for c in [cols["acc_proy"], cols["trx_proy"], cols["per_proy"]] if c)
            rows_cli.append({"Pais": current_pais or "", "Cliente": name_s,
                             "Monto USD": monto, "Real Excel": rt, "Proyectado Excel": pt})
    return pd.DataFrame(rows_cli), pd.DataFrame(rows_bol), {"engine":"xlcalculator","columns":cols}

def extract_cached(wb):
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    wbc = openpyxl.load_workbook(bio, data_only=True)
    ws = wbc[NEG_SHEET]; cols = locate_columns(ws); h = cols["header"]; ccli = cols["cliente"]
    bolsas_map = {"BCS":"Chile","BVL":"Perú","BVC":"Colombia"}
    rows_cli = []; rows_bol = []; current_pais = None
    for r in range(h+2, ws.max_row+1):
        name = ws.cell(r, ccli).value
        if name is None or str(name).strip()=="": continue
        name_s = str(name).strip()
        def coord(col): return ws.cell(r, col).coordinate if col else None
        if name_s in bolsas_map:
            current_pais = bolsas_map[name_s]
            rt = read_val(ws, coord(cols["real_total"] or cols["acc_real"] or cols["trx_real"] or cols["per_real"]))
            pt = read_val(ws, coord(cols["proy_total"] or cols["acc_proy"] or cols["trx_proy"] or cols["per_proy"]))
            rows_bol.append({"Pais": current_pais, "Bolsa": name_s, "Real Excel": rt, "Proyectado Excel": pt})
        else:
            monto = read_val(ws, coord(cols["monto"]))
            rt = read_val(ws, coord(cols["real_total"])) if cols["real_total"] else 0.0
            pt = read_val(ws, coord(cols["proy_total"])) if cols["proy_total"] else 0.0
            if rt == 0.0:
                rt = sum(read_val(ws, coord(c)) for c in [cols["acc_real"], cols["trx_real"], cols["per_real"]] if c)
            if pt == 0.0:
                pt = sum(read_val(ws, coord(c)) for c in [cols["acc_proy"], cols["trx_proy"], cols["per_proy"]] if c)
            rows_cli.append({"Pais": current_pais or "", "Cliente": name_s,
                             "Monto USD": monto, "Real Excel": rt, "Proyectado Excel": pt})
    return pd.DataFrame(rows_cli), pd.DataFrame(rows_bol), {"engine":"cached","columns":cols}

def recalc_and_extract(wb):
    # 1) xlcalculator
    if XL_OK:
        try:
            return extract_with_model(wb)
        except Exception as e:
            meta1 = {"engine":"xlcalculator_failed", "error": str(e)}
    else:
        meta1 = {"engine":"xlcalculator_unavailable"}
    # 2) cached values
    try:
        df1, df2, meta2 = extract_cached(wb)
        meta2.update(meta1); return df1, df2, meta2
    except Exception as e:
        meta2 = {"engine":"fallback", "error": str(e)}; meta2.update(meta1)
        return pd.DataFrame(columns=["Pais","Cliente","Monto USD","Real Excel","Proyectado Excel"]), \
               pd.DataFrame(columns=["Pais","Bolsa","Real Excel","Proyectado Excel"]), meta2

# -----------------------------
# UI
# -----------------------------
st.title("📊 Simulador Tarifario RV – Excel-driven (v8)")

uploaded = st.file_uploader("📁 Sube tu Excel maestro", type=["xlsx"])
if not uploaded:
    st.info("Se requieren las hojas '1. Parametros' y 'A.3 Negociación'.")
    st.stop()

raw_bytes = uploaded.read()
wb_prefill = load_workbook(raw_bytes)
params0 = read_params_from_wb(wb_prefill)

with st.sidebar:
    st.header("⚙️ Parámetros (columna R+)")
    desc_bcs = st.number_input("Descuento BCS (0–1)", min_value=0.0, max_value=1.0, step=0.01, value=float(params0.get("desc_bcs",0.15)))

    st.subheader("Transacción – Tramos")
    trans_edit = {}
    for pais in ["Chile","Colombia","Perú"]:
        base = params0["transaccion"].get(pais, [])
        df_e = pd.DataFrame(base) if base else pd.DataFrame(columns=["min","max","bps","fijo"])
        df_show = df_e.rename(columns={"min":"Mín (USD)","max":"Máx (USD)","bps":"Variable (fracción/%)","fijo":"Fijo (USD)"})
        df_show = st.data_editor(df_show, key=f"trx_{pais}", use_container_width=True, num_rows="dynamic")
        trans_edit[pais] = [{"min":safe_float(r.get("Mín (USD)",0)), "max":safe_float(r.get("Máx (USD)",float("inf"))),
                             "bps":safe_float(r.get("Variable (fracción/%)",0)), "fijo":safe_float(r.get("Fijo (USD)",0))}
                            for _, r in df_show.iterrows()]

    st.subheader("DMA – Tramos")
    dma_edit = {}
    for pais in ["Chile","Colombia","Perú"]:
        base = params0["dma"].get(pais, [])
        df_e = pd.DataFrame(base) if base else pd.DataFrame(columns=["min","max","bps","fijo"])
        df_show = df_e.rename(columns={"min":"Mín (USD)","max":"Máx (USD)","bps":"Variable (fracción/%)","fijo":"Fijo (USD)"})
        df_show = st.data_editor(df_show, key=f"dma_{pais}", use_container_width=True, num_rows="dynamic")
        dma_edit[pais] = [{"min":safe_float(r.get("Mín (USD)",0)), "max":safe_float(r.get("Máx (USD)",float("inf"))),
                           "bps":safe_float(r.get("Variable (fracción/%)",0)), "fijo":safe_float(r.get("Fijo (USD)",0))}
                          for _, r in df_show.iterrows()]

    st.subheader("Acceso – Códigos/Pantallas")
    pant_edit = {}
    for pais in ["Chile","Colombia","Perú"]:
        base = params0["pantallas"].get(pais, [])
        df_e = pd.DataFrame(base) if base else pd.DataFrame(columns=["min","max","var","fija"])
        df_show = df_e.rename(columns={"min":"Mín #Códigos","max":"Máx #Códigos","var":"Variable por código (USD)","fija":"Fijo mensual tramo (USD)"})
        df_show = st.data_editor(df_show, key=f"pant_{pais}", use_container_width=True, num_rows="dynamic")
        pant_edit[pais] = [{"min":safe_float(r.get("Mín #Códigos",0)), "max":safe_float(r.get("Máx #Códigos",float("inf"))),
                            "var":safe_float(r.get("Variable por código (USD)",0)), "fija":safe_float(r.get("Fijo mensual tramo (USD)",0))}
                           for _, r in df_show.iterrows()]

params_live = {"desc_bcs":desc_bcs, "transaccion":trans_edit, "dma":dma_edit, "pantallas":pant_edit}

with st.spinner("Aplicando parámetros y obteniendo resultados…"):
    wb = load_workbook(raw_bytes)
    wb = write_params_to_wb(wb, params_live)
    df_cli, df_bol, meta = recalc_and_extract(wb)

# KPIs y filtros
paises = ["Todos"] + sorted([p for p in df_cli["Pais"].dropna().unique().tolist() if isinstance(p,str) and p.strip()])
c1, c2 = st.columns(2)
with c1:
    pais_sel = st.selectbox("Filtrar por País", paises, index=0)
with c2:
    ver_detalle = st.toggle("Mostrar detalle por cliente", value=True)

df_f = df_cli if pais_sel=="Todos" else df_cli[df_cli["Pais"]==pais_sel]
tot_real = safe_float(df_f["Real Excel"].sum(),0.0)
tot_proy = safe_float(df_f["Proyectado Excel"].sum(),0.0)
monto    = safe_float(df_f["Monto USD"].sum(),0.0)

a,b,c,d = st.columns(4)
a.metric("Ingreso Real (Excel)", f"${tot_real:,.0f}")
b.metric("Ingreso Proyectado (Excel)", f"${tot_proy:,.0f}", delta=(f"{(tot_proy/tot_real-1)*100:+.1f}%" if tot_real>0 else None))
b.caption(f"Motor: **{meta.get('engine','?')}**")
c.metric("BPS Proyectado (Excel)", f"{calc_bps(tot_proy, monto):.2f} bps")
d.metric("Filas cargadas", f"{len(df_f):,}")

st.markdown("---")
cA, cB = st.columns(2, gap="large")
with cA:
    st.subheader("Real vs Proyectado – Excel")
    fig = go.Figure(data=[
        go.Bar(name="Real (Excel)", x=["Negociación"], y=[tot_real]),
        go.Bar(name="Proyectado (Excel)", x=["Negociación"], y=[tot_proy])
    ])
    fig.update_layout(barmode="group", height=300)
    st.plotly_chart(fig, use_container_width=True)
with cB:
    st.subheader("Totales por Bolsa (Excel)")
    st.dataframe(df_bol, use_container_width=True, height=300)

st.subheader("📋 Detalle por Cliente (Excel)")
if ver_detalle:
    tmp = df_f.copy()
    tmp["BPS"] = tmp.apply(lambda r: calc_bps(r["Proyectado Excel"], r["Monto USD"]), axis=1)
    st.dataframe(tmp.sort_values(["Pais","Proyectado Excel"], ascending=[True, False]), use_container_width=True, height=420)

st.markdown("---")
with st.expander("🔎 Diagnóstico"):
    st.json(meta)

col1, col2 = st.columns(2)
with col1:
    st.download_button("📥 Descargar Detalle (CSV)", df_f.to_csv(index=False).encode("utf-8"), "detalle_excel.csv", "text/csv", use_container_width=True)
with col2:
    st.download_button("📥 Descargar Excel (parametrizado, sin imágenes)",
                       safe_save_wb(remove_drawings_images(wb)), "excel_parametrizado.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
